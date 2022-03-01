import sys
from openpyxl import load_workbook
file_name = 'std_major.xlsx'
wb = load_workbook(filename = file_name, data_only = True)
ws = wb.active

# set hyperparameters
number = 65

# start iteration
for i in range(7):
    #row = chr(number)
    #row = 'A' + chr(number)
    row = 'B' + chr(number) # 이 때 range(7)
    #row = 'AZ'
    name = row + '1'
    st_num = 2
    fin_num = 33
    print(ws[name].value)

    # get information from xlsx file
    all_st = []
    for i in range(st_num,fin_num + 1):
        cell_num = row + str(i)
        if ws[cell_num].value == None : break
        print(ws[cell_num].value)
        all_st.append(ws[cell_num].value)

    # write text file
    f = open(ws[name].value + '.txt', 'w')
    f.write(str(len(all_st)) + '\n')
    for i in range(len(all_st)):
        f.write(all_st[i] + '\n')
    f.close()

    # go to next step
    number += 1
